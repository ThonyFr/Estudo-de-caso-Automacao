#Biblioteca para trabalhar com automação
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By

#Biblioteca para trabalhar com planilhas
from openpyxl import load_workbook

navegador = webdriver.Edge()

#Acessar um site
navegador.get('https://www.bing.com')

#Clica nos elementos indicados do site

navegador.find_element(By.XPATH, '//*[@id="sb_form_q"]').send_keys('Cotação do dólar')
navegador.find_element(By.XPATH, '//*[@id="sb_form_q"]').send_keys(Keys.ENTER)

cotacao_dolar = navegador.find_element(By.XPATH, '//*[@id="cc_tv"]').get_attribute('value')

print(f'Cotação do dólar atual {cotacao_dolar}')

arquivo = load_workbook()

#Retorna os nomes das planilhas
#print(arquivo.sheetnames)